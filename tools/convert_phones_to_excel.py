#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
医院电话表 Markdown → Excel 转换器
"""

import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 读取 Markdown 文件
def read_md_file(path):
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()

# 解析 Markdown 表格
def parse_md_table(table_text):
    lines = table_text.strip().split('\n')
    if len(lines) < 3:
        return []
    
    # 跳过表头分隔线 (|---|---|)
    data_lines = [l for l in lines if not re.match(r'^\|[\s\-:|]+\|$', l)]
    
    rows = []
    for line in data_lines:
        cells = [c.strip() for c in line.split('|')]
        # 移除首尾空单元格
        if cells and cells[0] == '':
            cells = cells[1:]
        if cells and cells[-1] == '':
            cells = cells[:-1]
        if cells:
            rows.append(cells)
    return rows

# 提取所有表格
def extract_tables(md_content):
    tables = []
    # 匹配 Markdown 表格
    pattern = r'(\|.*?\n\|[\s\-:|]+\n(?:\|.*?\n)+)'
    matches = re.findall(pattern, md_content, re.MULTILINE)
    
    for match in matches:
        rows = parse_md_table(match)
        if len(rows) >= 2:  # 至少表头 + 1 行数据
            tables.append(rows)
    
    return tables

# 提取章节标题
def extract_sections(md_content):
    sections = []
    # 匹配 ## 标题
    pattern = r'^##\s+(.+?)$'
    matches = re.finditer(pattern, md_content, re.MULTILINE)
    
    positions = []
    for m in matches:
        positions.append((m.start(), m.group(1).strip()))
    
    return positions

# 清理 sheet 名称（Excel 不允许 \/?:*[] 等字符）
def sanitize_sheet_name(name):
    invalid_chars = r'[]/\?:*"'
    for c in invalid_chars:
        name = name.replace(c, '')
    return name.strip()[:31]

# 创建 Excel
def create_excel(md_path, output_path):
    md_content = read_md_file(md_path)
    wb = Workbook()
    wb.remove(wb.active)  # 移除默认 sheet
    
    # 获取所有章节
    sections = extract_sections(md_content)
    
    # 按章节分割内容
    for i, (pos, section_title) in enumerate(sections):
        # 获取章节内容
        start = pos
        end = sections[i+1][0] if i+1 < len(sections) else len(md_content)
        section_content = md_content[start:end]
        
        # 提取表格
        tables = extract_tables(section_content)
        
        if tables:
            # 创建新 sheet
            sheet_name = sanitize_sheet_name(section_title)
            if not sheet_name:
                sheet_name = f'Sheet{i+1}'
            ws = wb.create_sheet(title=sheet_name)
            
            # 样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            row_offset = 0
            for table in tables:
                if len(table) < 2:
                    continue
                
                # 写入表头
                for col, header in enumerate(table[0], 1):
                    cell = ws.cell(row=row_offset+1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border
                
                # 写入数据
                for row_idx, row_data in enumerate(table[1:], 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_offset+row_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        cell.alignment = Alignment(horizontal='left')
                
                row_offset += len(table) + 1  # 表格间空一行
    
    # 保存
    wb.save(output_path)
    print(f"✅ 已保存：{output_path}")
    print(f"📊 共 {len(wb.sheetnames)} 个工作表")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")

if __name__ == '__main__':
    import sys
    md_path = sys.argv[1] if len(sys.argv) > 1 else '/home/admin/.openclaw/workspace/tools/hospital-phones.md'
    output_path = md_path.replace('.md', '.xlsx')
    create_excel(md_path, output_path)
